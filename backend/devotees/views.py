import csv
from django.http import HttpResponse, FileResponse
from rest_framework import generics, filters
from rest_framework.permissions import AllowAny
from django_filters.rest_framework import (
    DjangoFilterBackend,
    FilterSet,
    DateFromToRangeFilter,
    CharFilter,
    NumberFilter,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Devotee, Gothra, Nakshatra
from .serializers import (
    DevoteeSerializer,
    GothraSerializer,
    NakshatraSerializer,
    DevoteeFullSerializer,
)
from core.utils import TenantMixin
from .pdf_utils import generate_devotee_list_pdf


class DevoteeFilter(FilterSet):
    # Existing
    id = NumberFilter(field_name="id")
    full_name = CharFilter(field_name="full_name", lookup_expr="icontains")
    phone = CharFilter(field_name="phone", lookup_expr="icontains")
    email = CharFilter(field_name="email", lookup_expr="icontains")
    created_at = DateFromToRangeFilter(field_name="created_at")

    # Booking filters (may create duplicates without distinct)
    booking_date = DateFromToRangeFilter(field_name="bookings__booking_date")
    pooja = NumberFilter(field_name="bookings__pooja__id")  # expects an ID number

    # ✅ Added to match frontend filter: ?gothra=<id>
    gothra = NumberFilter(field_name="gothra__id")
    # ✅ Optional (good to have): ?nakshatra=<id>
    nakshatra = NumberFilter(field_name="nakshatra__id")

    class Meta:
        model = Devotee
        fields = [
            "full_name",
            "phone",
            "created_at",
            "booking_date",
            "pooja",
            "gothra",
            "nakshatra",
        ]


class DevoteeListCreateView(TenantMixin, generics.ListCreateAPIView):
    model = Devotee
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all().order_by("-id")
    serializer_class = DevoteeSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = DevoteeFilter
    search_fields = ["full_name", "phone", "email", "id_proof_number"]
    ordering_fields = ["id", "created_at", "full_name"]
    ordering = ["-id"]

    def get_queryset(self):
        return super().get_queryset().distinct()


class DevoteeDetailView(TenantMixin, generics.RetrieveUpdateDestroyAPIView):
    model = Devotee
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all()
    serializer_class = DevoteeSerializer


class GothraListCreateView(TenantMixin, generics.ListCreateAPIView):
    model = Gothra
    queryset = Gothra.objects.all().order_by("name")
    serializer_class = GothraSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name"]
    ordering = ["name"]

class GothraDetailView(TenantMixin, generics.RetrieveUpdateDestroyAPIView):
    model = Gothra
    queryset = Gothra.objects.all()
    serializer_class = GothraSerializer


class NakshatraListCreateView(TenantMixin, generics.ListCreateAPIView):
    model = Nakshatra
    queryset = Nakshatra.objects.all().order_by("name")
    serializer_class = NakshatraSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name"]
    ordering = ["name"]

class NakshatraDetailView(TenantMixin, generics.RetrieveUpdateDestroyAPIView):
    model = Nakshatra
    queryset = Nakshatra.objects.all()
    serializer_class = NakshatraSerializer


class DevoteeFullDetailView(TenantMixin, generics.RetrieveAPIView):
    model = Devotee
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all()
    serializer_class = DevoteeFullSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.prefetch_related("bookings", "donations")


class DevoteeExportCSVView(generics.GenericAPIView):
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all()
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="devotees.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "ID", "Full Name", "Phone", "Email", "Gothra", "Nakshatra",
            "ID Proof Type", "ID Proof Number", "Address", "Created At"
        ])

        for d in self.get_queryset().order_by("-id"):
            writer.writerow([
                d.id,
                d.full_name,
                d.phone,
                d.email,
                d.gothra.name if d.gothra else "",
                d.nakshatra.name if d.nakshatra else "",
                d.id_proof_type,
                d.id_proof_number,
                d.address,
                d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else "",
            ])

        return response


class DevoteeExportExcelView(generics.GenericAPIView):
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all()
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Devotees"

        headers = [
            "ID", "Full Name", "Phone", "Email", "Gothra", "Nakshatra",
            "ID Proof Type", "ID Proof Number", "Address", "Created At"
        ]
        ws.append(headers)

        for d in self.get_queryset().order_by("-id"):
            ws.append([
                d.id,
                d.full_name,
                d.phone,
                d.email,
                d.gothra.name if d.gothra else "",
                d.nakshatra.name if d.nakshatra else "",
                d.id_proof_type,
                d.id_proof_number,
                d.address,
                d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else "",
            ])

        for col_idx, col_name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(col_name) + 2)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="devotees.xlsx"'
        wb.save(response)
        return response


class DevoteeExportPDFView(generics.GenericAPIView):
    queryset = Devotee.objects.select_related("gothra", "nakshatra").all()
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        temple_name = getattr(request, "tenant", None)
        temple_name = temple_name.name if temple_name else "Temple"

        devotees = self.get_queryset().order_by("-id")
        buffer = generate_devotee_list_pdf(temple_name, devotees)

        return FileResponse(buffer, as_attachment=True, filename="devotees_list.pdf")
